from django.shortcuts import render
from django.db.models import Avg, Count, Q
from django.http import HttpResponse
from .models import Student, Subject, Result
import json, csv
import pandas as pd
from django.db.models import Avg, Count, Q, Case, When, IntegerField
PASS = 35
from django.http import HttpResponseBadRequest
from django.contrib.auth.decorators import login_required, user_passes_test

# ================= DASHBOARD =================
from django.shortcuts import render
from django.db.models import Avg, Count, Q
from django.http import HttpResponse
from .models import Student, Subject, Result
import json, csv
import pandas as pd

PASS = 35


from django.shortcuts import render
from django.db.models import Avg, Count, Q
from .models import Result
import json

PASS = 35

@login_required
def dashboard(request):
    qs = Result.objects.select_related('student', 'subject')

    # ===== FILTERS =====
    year = request.GET.get('year')
    sem = request.GET.get('sem')
    branch = request.GET.get('branch')
    name = request.GET.get('name')

    if year:
        qs = qs.filter(year=year)
    if sem:
        qs = qs.filter(semester=sem)
    if branch:
        qs = qs.filter(student__branch=branch)
    if name:
        qs = qs.filter(student__name__icontains=name)

    # ===== KPI =====
    total = qs.count()
    pass_count = qs.filter(marks__gte=PASS).count()
    fail_count = qs.filter(marks__lt=PASS).count()

    pass_pct = round((pass_count / total) * 100, 2) if total else 0
    avg_marks = round(qs.aggregate(avg=Avg('marks'))['avg'] or 0, 2)

    # ===== DISTINCTION / FIRST CLASS =====
    distinction = qs.filter(marks__gte=75)\
        .values('student').distinct().count()

    first_class = qs.filter(marks__gte=60, marks__lt=75)\
        .values('student').distinct().count()

    # ===== CHART DATA =====
    trend = list(
        qs.values('semester')
        .annotate(avg=Avg('marks'))
        .order_by('semester')
    )

    subject_difficulty = list(
        qs.values('subject__code')
        .annotate(fail_count=Count('id', filter=Q(marks__lt=PASS)))
    )

    sgpa_list = list(
        qs.values('student__name')
        .annotate(avg=Avg('marks'))
        .order_by('-avg')[:5]
    )

    branch_data = list(
        qs.values('student__branch')
        .annotate(avg=Avg('marks'))
    )

    category_data = list(
        qs.values('student__actual_category')
        .annotate(count=Count('id'))
    )

    # ===== STUDENTS =====
    students = qs.values(
        'student__usn',
        'student__name',
        'student__branch',
        'student__actual_category',
        'student__admission_quota'
    ).annotate(
        avg=Avg('marks'),
        backlog=Count('id', filter=Q(marks__lt=PASS))
    ).order_by('-avg')

    # ✅ ===== ADD THIS BLOCK (THIS WAS MISSING) =====
    years = Result.objects.values_list('year', flat=True).distinct()
    sems = Result.objects.values_list('semester', flat=True).distinct()
    branches = Result.objects.values_list('student__branch', flat=True).distinct()

    # ===== RETURN =====
    return render(request, 'analytics/dashboard.html', {
        'total': total,
        'pass_pct': pass_pct,
        'avg': avg_marks,
        'backlog_count': fail_count,

        'distinction': distinction,
        'first_class': first_class,

        'trend': json.dumps(trend),
        'subject_difficulty': json.dumps(subject_difficulty),
        'sgpa_list': json.dumps(sgpa_list),
        'branch_data': json.dumps(branch_data),
        'category_data': json.dumps(category_data),

        'students': students,

        # ✅ ADD THESE
        'years': years,
        'sems': sems,
        'branches': branches,
    })



from django.shortcuts import render
from django.db.models import Avg, Count, Q
from .models import Result
import json

PASS = 35

from django.shortcuts import render
from django.db.models import Avg, Count, Q, Case, When, IntegerField
from .models import Result
import json

PASS = 35

def get_subject_analysis(qs):
    subjects = qs.values('subject__code', 'subject__name').annotate(
        total=Count('id'),
        passed=Count('id', filter=Q(marks__gte=35)),
        failed=Count('id', filter=Q(marks__lt=35)),

        c50=Count('id', filter=Q(marks__gte=50, marks__lt=60)),
        c60=Count('id', filter=Q(marks__gte=60, marks__lt=70)),
        c70=Count('id', filter=Q(marks__gte=70)),
    )

    data = []
    for s in subjects:
        total = s['total']
        pass_pct = round((s['passed']/total)*100,2) if total else 0
        fail_pct = round((s['failed']/total)*100,2) if total else 0

        data.append({
            'code': s['subject__code'],
            'name': s['subject__name'],
            'total': total,
            'passed': s['passed'],
            'failed': s['failed'],
            'pass_pct': pass_pct,
            'fail_pct': fail_pct,
            'c50': s['c50'],
            'c60': s['c60'],
            'c70': s['c70'],
        })

    # IMPORTANT → sort for clean charts
    return sorted(data, key=lambda x: x['pass_pct'], reverse=True)

def subject_view(request):
    qs = Result.objects.select_related('student', 'subject')

    year = request.GET.get('year')
    sem = request.GET.get('sem')
    branch = request.GET.get('branch')
    subject = request.GET.get('subject')

    if year:
        qs = qs.filter(year=year)
    if sem:
        qs = qs.filter(semester=sem)
    if branch:
        qs = qs.filter(student__branch=branch)
    if subject:
        qs = qs.filter(subject__code=subject)

    # ✅ ADDED CLEAN ANALYSIS
    analysis = get_subject_analysis(qs)

    # ===== DROPDOWNS =====
    years = Result.objects.values_list('year', flat=True).distinct()
    sems = Result.objects.values_list('semester', flat=True).distinct()
    branches = Result.objects.values_list('student__branch', flat=True).distinct()
    subjects = Result.objects.values_list('subject__code', flat=True).distinct()

    # ===== KPI =====
    total_students = qs.values('student').distinct().count()
    pass_count = qs.filter(marks__gte=PASS).count()
    total_records = qs.count()

    pass_pct = round((pass_count / total_records * 100), 2) if total_records else 0
    avg_score = round(qs.aggregate(avg=Avg('marks'))['avg'] or 0, 2)
    fail_count = qs.filter(marks__lt=PASS).count()

    # ===== SUBJECT ANALYSIS (OLD CHART DATA - KEEP) =====
    data = qs.values('subject__code').annotate(
        avg=Avg('marks'),
        pass_count=Count('id', filter=Q(marks__gte=PASS)),
        fail_count=Count('id', filter=Q(marks__lt=PASS)),
        total=Count('id')
    )

    labels, avg_marks, pass_rates, fail_rates = [], [], [], []
    heat_labels, heat_values = [], []

    for d in data:
        labels.append(d['subject__code'])
        avg_marks.append(round(d['avg'] or 0, 2))

        total = d['total']
        pass_rates.append(round((d['pass_count']/total)*100 if total else 0, 2))
        fail_rates.append(round((d['fail_count']/total)*100 if total else 0, 2))

        fail_rate = (d['fail_count']/total)*100 if total else 0
        heat_labels.append(d['subject__code'])
        heat_values.append(round(fail_rate, 1))

    # ===== HARDEST =====
    hardest = sorted(zip(labels, fail_rates), key=lambda x: x[1], reverse=True)[:5]

    # ===== GRADE DISTRIBUTION =====
    grade_qs = qs.annotate(
        grade_calc=Case(
            When(marks__gte=75, then=1),
            When(marks__gte=60, then=2),
            When(marks__gte=35, then=3),
            default=4,
            output_field=IntegerField()
        )
    ).values('grade_calc').annotate(count=Count('id'))

    grade_map = {1:'Distinction',2:'First Class',3:'Pass',4:'Fail'}
    grade_labels, grade_counts = [], []

    for g in grade_qs:
        grade_labels.append(grade_map[g['grade_calc']])
        grade_counts.append(g['count'])

    # ===== STUDENTS =====
    students = qs.values(
        'student__usn',
        'student__name',
        'student__branch',
        'student__actual_category',
        'student__admission_quota'
    ).annotate(avg=Avg('marks')).order_by('-avg')

    return render(request, 'analytics/subject.html', {
        'years': years,
        'sems': sems,
        'branches': branches,
        'subjects': subjects,

        'total_students': total_students,
        'pass_pct': pass_pct,
        'avg_score': avg_score,
        'fail_count': fail_count,

        'labels': json.dumps(labels),
        'avg_marks': json.dumps(avg_marks),
        'pass_rates': json.dumps(pass_rates),
        'fail_rates': json.dumps(fail_rates),

        'hard_labels': json.dumps([x[0] for x in hardest]),
        'hard_data': json.dumps([x[1] for x in hardest]),

        'grade_labels': json.dumps(grade_labels),
        'grade_counts': json.dumps(grade_counts),

        'heat_labels': heat_labels,
        'heat_values': heat_values,

        'students': students,

        # ✅ NEW CLEAN DATA
        'analysis': analysis
    })
def category(request):
    from django.db.models import Count, Avg, Q
    import json

    PASS = 35
    qs = Result.objects.select_related('student', 'subject')

    # ===== FILTERS =====
    year = request.GET.get('year')
    sem = request.GET.get('sem')
    branch = request.GET.get('branch')
    subject_filter = request.GET.get('subject')

    if year:
        qs = qs.filter(year=year)
    if sem:
        qs = qs.filter(semester=sem)
    if branch:
        qs = qs.filter(student__branch=branch)
    if subject_filter:
        qs = qs.filter(subject__code=subject_filter)

    # ===== DROPDOWNS =====
    years = Result.objects.values_list('year', flat=True).distinct()
    sems = Result.objects.values_list('semester', flat=True).distinct()
    branches = Result.objects.values_list('student__branch', flat=True).distinct()
    subjects = Result.objects.values_list('subject__code', flat=True).distinct()

    # ===== CATEGORY DISTRIBUTION =====
    dist = qs.values('student__actual_category').annotate(count=Count('id'))

    labels = [d['student__actual_category'] or "Unknown" for d in dist]
    counts = [d['count'] for d in dist]

    # ===== AVG + PASS + BACKLOG =====
    perf = qs.values('student__actual_category').annotate(
        total=Count('id'),
        passed=Count('id', filter=Q(marks__gte=PASS)),
        avg=Avg('marks')
    )

    avg_scores = []
    pass_rates = []
    backlog_rates = []

    for d in perf:
        total = d['total']
        passed = d['passed']
        failed = total - passed

        avg_scores.append(round(d['avg'] or 0, 2))
        pass_rates.append(round((passed / total) * 100, 2) if total else 0)
        backlog_rates.append(round((failed / total) * 100, 2) if total else 0)

    # ===== STUDENTS TABLE =====
    students = qs.values(
        'student__usn',
        'student__name',
        'student__branch',
        'student__actual_category',
        'student__admission_quota'
    ).annotate(
        avg=Avg('marks'),
        backlog=Count('id', filter=Q(marks__lt=PASS))
    ).order_by('-avg')

    return render(request, 'analytics/category.html', {
        'years': years,
        'sems': sems,
        'branches': branches,
        'subjects': subjects,

        'labels': json.dumps(labels),
        'counts': json.dumps(counts),
        'avg_scores': json.dumps(avg_scores),
        'pass_rates': json.dumps(pass_rates),
        'backlog_rates': json.dumps(backlog_rates),

        'students': students
    })

def branch(request):
    from django.db.models import Count, Avg, Q
    import json

    PASS = 35
    qs = Result.objects.select_related('student')

    year = request.GET.get('year')
    sem = request.GET.get('sem')

    if year:
        qs = qs.filter(year=year)
    if sem:
        qs = qs.filter(semester=sem)

    # ===== BRANCH LIST =====
    branches = list(qs.values_list('student__branch', flat=True).distinct())

    # ===== DISTRIBUTION =====
    dist = qs.values('student__branch').annotate(
        count=Count('student', distinct=True)
    )

    labels = [d['student__branch'] for d in dist]
    counts = [d['count'] for d in dist]

    # ===== PERFORMANCE =====
    perf = qs.values('student__branch').annotate(
        total=Count('id'),
        passed=Count('id', filter=Q(marks__gte=PASS)),
        avg=Avg('marks')
    )

    pass_rates = []
    avg_scores = []

    for d in perf:
        total = d['total']
        passed = d['passed']

        pass_rates.append(round((passed/total)*100,2) if total else 0)
        avg_scores.append(round(d['avg'] or 0,2))

    # ===== PASS FAIL =====
    pf = qs.values('student__branch').annotate(
        pass_count=Count('id', filter=Q(marks__gte=PASS)),
        fail_count=Count('id', filter=Q(marks__lt=PASS))
    )

    pf_pass = [d['pass_count'] for d in pf]
    pf_fail = [d['fail_count'] for d in pf]

    # ===== TOPPER =====
    top_labels = []
    top_marks = []

    for b in branches:
        top = qs.filter(student__branch=b).order_by('-marks').first()
        if top:
            top_labels.append(b)
            top_marks.append(top.marks)

    # ===== GRADE =====
    grade = qs.values('student__branch').annotate(
        below40=Count('id', filter=Q(marks__lt=40)),
        between40_75=Count('id', filter=Q(marks__gte=40, marks__lt=75)),
        above75=Count('id', filter=Q(marks__gte=75))
    )

    g1 = [d['below40'] for d in grade]
    g2 = [d['between40_75'] for d in grade]
    g3 = [d['above75'] for d in grade]

    # ===== TABLE =====
    table = []

    for b in branches:
        bqs = qs.filter(student__branch=b)

        students = bqs.values('student').distinct().count()
        avg = round(bqs.aggregate(avg=Avg('marks'))['avg'] or 0,2)

        total = bqs.count()
        passed = bqs.filter(marks__gte=PASS).count()
        failed = total - passed

        pass_rate = round((passed/total)*100,2) if total else 0

        top = bqs.order_by('-marks').first()
        topper = f"{top.student.name} ({top.marks})" if top else "-"

        distinction = bqs.filter(marks__gte=75).count()

        table.append({
            'branch': b,
            'students': students,
            'avg': avg,
            'pass_rate': pass_rate,
            'backlog': failed,
            'distinction': distinction,
            'topper': topper
        })
    years = Result.objects.values_list('year', flat=True).distinct()
    sems = Result.objects.values_list('semester', flat=True).distinct()

    return render(request, 'analytics/branch.html', {
        'labels': json.dumps(labels),
        'counts': json.dumps(counts),

        'pass_rates': json.dumps(pass_rates),
        'avg_scores': json.dumps(avg_scores),

        'pf_pass': json.dumps(pf_pass),
        'pf_fail': json.dumps(pf_fail),

        'top_labels': json.dumps(top_labels),
        'top_marks': json.dumps(top_marks),

        'g1': json.dumps(g1),
        'g2': json.dumps(g2),
        'g3': json.dumps(g3),

        'table': table,
        'years': years,
        'sems': sems,
    })


# ================= BACKLOG =================
def backlog(request):
    from django.db.models import Count, Avg, Q, Max
    import json

    PASS = 35
    qs = Result.objects.select_related('student', 'subject')

    # ===== FILTERS =====
    year = request.GET.get('year')
    sem = request.GET.get('sem')
    branch = request.GET.get('branch')
    subject = request.GET.get('subject')

    if year:
        qs = qs.filter(year=year)
    if sem:
        qs = qs.filter(semester=sem)
    if branch:
        qs = qs.filter(student__branch=branch)
    if subject:
        qs = qs.filter(subject__code=subject)

    # ===== STUDENT BACKLOG COUNT =====
    student_backlogs = qs.values('student').annotate(
        backlog_count=Count('id', filter=Q(marks__lt=PASS))
    )

    # ===== KPIs =====
    total_backlogs = qs.filter(marks__lt=PASS).count()

    students_with_backlog = student_backlogs.filter(
        backlog_count__gt=0
    ).count()

    max_backlog = student_backlogs.aggregate(
        Max('backlog_count')
    )['backlog_count__max'] or 0

    avg_backlog = round(
        student_backlogs.filter(backlog_count__gt=0)
        .aggregate(avg=Avg('backlog_count'))['avg'] or 0, 2
    )

    # ===== DISTRIBUTION =====
    dist_map = {i: 0 for i in range(1, 6)}
    for s in student_backlogs:
        c = s['backlog_count']
        if c >= 1:
            dist_map[min(c, 5)] += 1

    dist_labels = ['1 backlog','2 backlogs','3 backlogs','4 backlogs','5+']
    dist_data = list(dist_map.values())

    # ===== BRANCH BACKLOG RATE =====
    branch_data = qs.values('student__branch').annotate(
        total=Count('student', distinct=True),
        backlog_students=Count('student', filter=Q(marks__lt=PASS), distinct=True)
    )

    branch_labels = []
    branch_rates = []

    for b in branch_data:
        branch_labels.append(b['student__branch'])
        rate = (b['backlog_students'] / b['total'] * 100) if b['total'] else 0
        branch_rates.append(round(rate, 2))

    # ===== RISK =====
    safe = student_backlogs.filter(backlog_count=0).count()
    low = student_backlogs.filter(backlog_count=1).count()
    medium = student_backlogs.filter(backlog_count__in=[2,3]).count()
    high = student_backlogs.filter(backlog_count__gte=4).count()

    risk_labels = ['Safe','Low','Medium','High']
    risk_data = [safe, low, medium, high]

    # ===== SUBJECT FAIL =====
    subject_fail = qs.filter(marks__lt=PASS).values('subject__name').annotate(
        count=Count('id')
    ).order_by('-count')[:10]

    sub_labels = [s['subject__name'] for s in subject_fail]
    sub_data = [s['count'] for s in subject_fail]

    # ===== STUDENT TABLE =====
    students = qs.values(
        'student__usn','student__name','student__branch',
        'student__actual_category','student__admission_quota'
    ).annotate(
        avg=Avg('marks'),
        backlog=Count('id', filter=Q(marks__lt=PASS))
    ).filter(backlog__gt=0).order_by('-backlog')[:50]

    # ===== DROPDOWN DATA (THIS FIXES YOUR ISSUE) =====
    branches = Result.objects.values_list('student__branch', flat=True).distinct()
    subjects = Result.objects.values_list('subject__code', flat=True).distinct()
    years = Result.objects.values_list('year', flat=True).distinct()
    sems = Result.objects.values_list('semester', flat=True).distinct()

    return render(request, 'analytics/backlog.html', {

        'total_backlogs': total_backlogs,
        'students_with_backlog': students_with_backlog,
        'max_backlog': max_backlog,
        'avg_backlog': avg_backlog,

        'dist_labels': json.dumps(dist_labels),
        'dist_data': json.dumps(dist_data),

        'branch_labels': json.dumps(branch_labels),
        'branch_rates': json.dumps(branch_rates),

        'risk_labels': json.dumps(risk_labels),
        'risk_data': json.dumps(risk_data),

        'sub_labels': json.dumps(sub_labels),
        'sub_data': json.dumps(sub_data),

        'students': students,

        # ✅ THIS WAS MISSING
        'branches': branches,
        'subjects': subjects,
        'years': years,
        'sems': sems,
    })
def is_faculty_or_admin(user):
    return user.is_authenticated and (user.is_staff or user.is_superuser)

# ================= UPLOAD =================
@login_required
@user_passes_test(is_faculty_or_admin)
def upload(request):
    import pandas as pd
    import csv
    from .models import Student, Subject, Result, Backlog

    context = {}

    if request.method == 'POST' and request.FILES.get('file'):
        file = request.FILES['file']

        try:
            # =========================
            # READ FILE (CSV / EXCEL)
            # =========================
            if file.name.endswith(('.xlsx', '.xls')):
                df = pd.read_excel(file)
                df.columns = [c.strip().lower().replace(" ", "_") for c in df.columns]
                reader = df.to_dict(orient='records')
            else:
                decoded = file.read().decode('utf-8-sig').splitlines()
                reader = csv.DictReader(decoded)

            accepted, rejected = 0, 0
            errors = []

            # =========================
            # PROCESS ROWS
            # =========================
            for i, row in enumerate(reader, start=2):
                try:
                    if not row:
                        continue

                    # NORMALIZE KEYS + VALUES
                    row = {
                        k.strip().lower().replace(" ", "_"): (str(v).strip() if v else "")
                        for k, v in row.items()
                    }

                    usn = row.get('usn', '')
                    name = row.get('name', '')
                    branch = row.get('branch', '')
                    sem = row.get('semester', '')
                    year = row.get('year', '')
                    category = row.get('category', '')
                    quota = row.get('admission_quota', '')
                    sub_code = row.get('subject_code', '')
                    sub_name = row.get('subject_name', '')
                    marks = row.get('marks', '')
                    sgpa = row.get('sgpa', '')

                    # =========================
                    # VALIDATION
                    # =========================
                    if usn == "" or sub_code == "":
                        rejected += 1
                        errors.append(i)
                        continue

                    if marks == "" or sgpa == "":
                        rejected += 1
                        errors.append(i)
                        continue

                    try:
                        marks = float(marks)
                        sgpa = float(sgpa)
                    except:
                        rejected += 1
                        errors.append(i)
                        continue

                    if not (0 <= marks <= 100) or not (0 <= sgpa <= 10):
                        rejected += 1
                        errors.append(i)
                        continue

                    # =========================
                    # SAVE STUDENT
                    # =========================
                    student, _ = Student.objects.update_or_create(
                        usn=usn,
                        defaults={
                            'name': name,
                            'branch': branch,
                            'actual_category': category,
                            'admission_quota': quota,
                            'batch_year': int(year) if year else 2023
                        }
                    )

                    # =========================
                    # SAVE SUBJECT
                    # =========================
                    subject, _ = Subject.objects.get_or_create(
                        code=sub_code,
                        defaults={
                            'name': sub_name,
                            'branch': branch if branch else "CSE",
                            'semester': int(sem) if sem else 1
                        }
                    )

                    # =========================
                    # SAVE RESULT
                    # =========================
                    result, _ = Result.objects.update_or_create(
                        student=student,
                        subject=subject,
                        semester=int(sem) if sem else 1,
                        defaults={
                            'marks': marks,
                            'sgpa': sgpa,
                            'year': int(year) if year else 2023
                        }
                    )

                    # =========================
                    # 🔥 SAVE BACKLOG
                    # =========================
                    if marks < 35:
                        Backlog.objects.get_or_create(
                            student=student,
                            subject=subject,
                            semester=int(sem) if sem else 1,
                            year=int(year) if year else 2023,
                            defaults={"cleared": False}
                        )

                    accepted += 1

                except Exception:
                    rejected += 1
                    errors.append(i)

            context = {
                'accepted': accepted,
                'rejected': rejected,
                'errors': errors[:10]
            }

        except Exception as e:
            context = {
                'accepted': 0,
                'rejected': 0,
                'errors': [str(e)]
            }

    return render(request, 'analytics/upload.html', context)


# ================= QUOTA =================
def quota(request):
    from django.db.models import Count, Avg, Q
    import json

    PASS = 35
    qs = Result.objects.select_related('student', 'subject')

    # ===== FILTERS =====
    year = request.GET.get('year')
    sem = request.GET.get('sem')
    branch = request.GET.get('branch')
    subject = request.GET.get('subject')

    if year:
        qs = qs.filter(year=year)

    if sem:
        qs = qs.filter(semester=sem)

    if branch:
        qs = qs.filter(student__branch=branch)

    if subject:
        qs = qs.filter(subject__code=subject)   # change to subject__name if needed

    # ===== QUOTA DISTRIBUTION =====
    dist = qs.values('student__admission_quota').annotate(
        count=Count('student', distinct=True)
    )

    q_labels = [d['student__admission_quota'] or "Unknown" for d in dist]
    q_counts = [d['count'] for d in dist]

    # ===== PERFORMANCE =====
    perf = qs.values('student__admission_quota').annotate(
        total=Count('id'),
        passed=Count('id', filter=Q(marks__gte=PASS)),
        avg=Avg('marks'),
        distinction=Count('id', filter=Q(marks__gte=75))
    )

    avg_scores = []
    pass_rates = []
    backlog_rates = []
    distinctions = []
    table = []

    for d in perf:
        total = d['total']
        passed = d['passed']
        failed = total - passed

        avg = round(d['avg'] or 0, 2)
        pass_rate = round((passed/total)*100, 2) if total else 0
        backlog = round((failed/total)*100, 2) if total else 0

        avg_scores.append(avg)
        pass_rates.append(pass_rate)
        backlog_rates.append(backlog)
        distinctions.append(d['distinction'])

        table.append({
            'quota': d['student__admission_quota'] or "Unknown",
            'students': total,
            'avg': avg,
            'pass_rate': pass_rate,
            'backlog': backlog,
            'distinction': d['distinction']
        })

    # ===== STUDENTS =====
    students = qs.values(
        'student__usn',
        'student__name',
        'student__branch',
        'student__actual_category',
        'student__admission_quota'
    ).annotate(
        avg=Avg('marks'),
        backlog=Count('id', filter=Q(marks__lt=PASS))
    ).order_by('-avg')[:100]

    # ===== DROPDOWNS =====
    branches = Result.objects.values_list('student__branch', flat=True).distinct()
    subjects = Result.objects.values_list('subject__code', flat=True).distinct()
    years = Result.objects.values_list('year', flat=True).distinct()
    sems = Result.objects.values_list('semester', flat=True).distinct()

    return render(request, 'analytics/quota.html', {
        'q_labels': json.dumps(q_labels),
        'q_counts': json.dumps(q_counts),

        'avg_scores': json.dumps(avg_scores),
        'pass_rates': json.dumps(pass_rates),
        'backlog_rates': json.dumps(backlog_rates),

        'table': table,
        'students': students,

        'branches': branches,
        'subjects': subjects,
        'years': years,
        'sems': sems,
    })

    
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from django.db.models import Avg

def download_dashboard_report(request):
    from django.http import HttpResponse
    from openpyxl import Workbook
    from openpyxl.chart import BarChart, PieChart, LineChart, Reference
    from django.db.models import Avg

    PASS = 35

    qs = Result.objects.select_related('student')

    year = request.GET.get('year')
    sem = request.GET.get('sem')
    branch = request.GET.get('branch')

    if year:
        qs = qs.filter(year=year)
    if sem:
        qs = qs.filter(semester=sem)
    if branch:
        qs = qs.filter(student__branch=branch)

    wb = Workbook()
    ws = wb.active
    ws.title = "Dashboard"

    # =========================
    # 1️⃣ PASS / FAIL
    # =========================
    total = qs.count()
    passed = qs.filter(marks__gte=PASS).count()
    failed = qs.filter(marks__lt=PASS).count()

    ws.append(["Metric","Count"])
    ws.append(["Pass", passed])
    ws.append(["Fail", failed])

    pie = PieChart()
    pie.title = "Pass vs Fail"

    data = Reference(ws, min_col=2, min_row=1, max_row=3)
    cats = Reference(ws, min_col=1, min_row=2, max_row=3)

    pie.add_data(data, titles_from_data=True)
    pie.set_categories(cats)

    ws.add_chart(pie, "E2")

    # =========================
    # 2️⃣ SCORE DISTRIBUTION
    # =========================
    ws.append([])
    ws.append(["Range","Students"])

    ranges = {
        "<50": qs.filter(marks__lt=50).count(),
        "50-60": qs.filter(marks__gte=50, marks__lt=60).count(),
        "60-70": qs.filter(marks__gte=60, marks__lt=70).count(),
        "70-85": qs.filter(marks__gte=70, marks__lt=85).count(),
        ">=85": qs.filter(marks__gte=85).count(),
    }

    start = ws.max_row

    for k,v in ranges.items():
        ws.append([k,v])

    end = ws.max_row

    dist_chart = BarChart()
    dist_chart.title = "Score Distribution"

    data = Reference(ws, min_col=2, min_row=start, max_row=end)
    cats = Reference(ws, min_col=1, min_row=start+1, max_row=end)

    dist_chart.add_data(data, titles_from_data=True)
    dist_chart.set_categories(cats)

    for s in dist_chart.series:
        s.data_labels = True

    ws.add_chart(dist_chart, "E20")

    # =========================
    # 3️⃣ RESULT BREAKDOWN
    # =========================
    ws.append([])
    ws.append(["Result Type","Count"])

    distinction = qs.filter(marks__gte=75).count()
    first = qs.filter(marks__gte=60, marks__lt=75).count()
    pass_cls = qs.filter(marks__gte=35, marks__lt=60).count()
    fail = qs.filter(marks__lt=35).count()

    start = ws.max_row

    ws.append(["Distinction", distinction])
    ws.append(["First Class", first])
    ws.append(["Pass", pass_cls])
    ws.append(["Fail", fail])

    end = ws.max_row

    result_chart = BarChart()
    result_chart.title = "Result Breakdown"

    data = Reference(ws, min_col=2, min_row=start, max_row=end)
    cats = Reference(ws, min_col=1, min_row=start+1, max_row=end)

    result_chart.add_data(data, titles_from_data=True)
    result_chart.set_categories(cats)

    for s in result_chart.series:
        s.data_labels = True

    ws.add_chart(result_chart, "E38")

    # =========================
    # 4️⃣ 🔥 SGPA TREND (REAL DATA)
    # =========================
    ws.append([])
    ws.append(["Year-Sem","Avg SGPA"])

    trend = qs.values('year','semester').annotate(
        avg_sgpa=Avg('sgpa')   # ✅ using user data
    ).order_by('year','semester')

    start = ws.max_row

    for t in trend:
        label = f"{t['year']}-S{t['semester']}"
        ws.append([label, round(t['avg_sgpa'] or 0, 2)])

    end = ws.max_row

    if end > start:
        sgpa_chart = LineChart()
        sgpa_chart.title = "SGPA Trend"

        data = Reference(ws, min_col=2, min_row=start, max_row=end)
        cats = Reference(ws, min_col=1, min_row=start+1, max_row=end)

        sgpa_chart.add_data(data, titles_from_data=True)
        sgpa_chart.set_categories(cats)

        sgpa_chart.y_axis.title = "SGPA"
        sgpa_chart.x_axis.title = "Year-Sem"

        ws.add_chart(sgpa_chart, "E55")

    # =========================
    # DOWNLOAD
    # =========================
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=dashboard_report.xlsx'

    wb.save(response)
    return response

def download_subject_report(request):

    qs = Result.objects.select_related('student','subject')

    year = request.GET.get('year')
    sem = request.GET.get('sem')

    if year:
        qs = qs.filter(year=year)
    if sem:
        qs = qs.filter(semester=sem)

    wb = Workbook()
    ws = wb.active
    ws.title = "Subject Analysis"

    # ===== HEADER =====
    ws.append([
        "Subject",
        "Total",
        "Passed",
        "Failed",
        "Pass %"
    ])

    subjects = qs.values('subject__code').distinct()

    chart_data = []

    for s in subjects:
        sub_qs = qs.filter(subject__code=s['subject__code'])

        total = sub_qs.count()
        passed = sub_qs.filter(marks__gte=35).count()
        failed = sub_qs.filter(marks__lt=35).count()

        pass_pct = round((passed/total)*100,2) if total else 0

        ws.append([
            s['subject__code'],
            total,
            passed,
            failed,
            pass_pct
        ])

    # ===== CHART =====
    chart = BarChart()
    chart.title = "Subject Pass Percentage"
    chart.y_axis.title = "Pass %"
    chart.x_axis.title = "Subjects"

    data = Reference(ws, min_col=5, min_row=1, max_row=ws.max_row)
    categories = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)

    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)

    ws.add_chart(chart, "G2")

    # ===== DOWNLOAD =====
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=subject_report.xlsx'

    wb.save(response)
    return response

from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from django.db.models import Count, Q

def download_subject_report(request):
    from django.http import HttpResponse
    from openpyxl import Workbook
    from openpyxl.chart import BarChart, Reference
    from django.db.models import Count, Q

    PASS = 35

    qs = Result.objects.select_related('student','subject')

    # ===== FILTERS =====
    year = request.GET.get('year')
    sem = request.GET.get('sem')
    branch = request.GET.get('branch')

    if year and year != "All Years":
        qs = qs.filter(year=year)

    if sem and sem != "All Sem":
        qs = qs.filter(semester=sem)

    if branch and branch != "All Branch":
        qs = qs.filter(student__branch=branch)

    # ===== EXCEL =====
    wb = Workbook()
    ws = wb.active
    ws.title = "Subject Analysis"

    # FILTER INFO
    ws.append([f"Year: {year or 'All'}", f"Sem: {sem or 'All'}", f"Branch: {branch or 'All'}"])
    ws.append([])

    # HEADER
    ws.append([
        "Subject Code","Subject Name","Total",
        "Pass %","Fail %",
        "50-59","60-69",">=70"
    ])

    subjects = qs.values('subject__code','subject__name').distinct()

    for s in subjects:
        sub = qs.filter(subject__code=s['subject__code'])

        total = sub.count()
        passed = sub.filter(marks__gte=35).count()
        failed = sub.filter(marks__lt=35).count()

        pass_pct = round((passed/total)*100,2) if total else 0
        fail_pct = round((failed/total)*100,2) if total else 0

        c50 = sub.filter(marks__gte=50, marks__lt=60).count()
        c60 = sub.filter(marks__gte=60, marks__lt=70).count()
        c70 = sub.filter(marks__gte=70).count()

        ws.append([
            s['subject__code'],
            s['subject__name'],
            total,
            pass_pct,
            fail_pct,
            c50,
            c60,
            c70
        ])

    last = ws.max_row
    cats = Reference(ws, min_col=1, min_row=4, max_row=last)

    # =====================================
    # 1️⃣ PASS vs FAIL CHART
    # =====================================
    pf_chart = BarChart()
    pf_chart.title = "Pass % vs Fail %"
    pf_chart.y_axis.title = "%"
    pf_chart.x_axis.title = "Subjects"

    data = Reference(ws, min_col=4, max_col=5, min_row=3, max_row=last)
    pf_chart.add_data(data, titles_from_data=True)
    pf_chart.set_categories(cats)

    for s in pf_chart.series:
        s.data_labels = True

    ws.add_chart(pf_chart, "K3")

    # =====================================
    # 2️⃣ SCORE DISTRIBUTION
    # =====================================
    dist_chart = BarChart()
    dist_chart.title = "Score Distribution"
    dist_chart.x_axis.title = "Subjects"

    data = Reference(ws, min_col=6, max_col=8, min_row=3, max_row=last)
    dist_chart.add_data(data, titles_from_data=True)
    dist_chart.set_categories(cats)

    dist_chart.gapWidth = 150

    for s in dist_chart.series:
        s.data_labels = True

    ws.add_chart(dist_chart, "K20")

    # =====================================
    # 3️⃣ PASS % ONLY (CLEAR VIEW)
    # =====================================
    pass_chart = BarChart()
    pass_chart.title = "Pass % by Subject"

    data = Reference(ws, min_col=4, min_row=3, max_row=last)
    pass_chart.add_data(data, titles_from_data=True)
    pass_chart.set_categories(cats)

    for s in pass_chart.series:
        s.data_labels = True

    ws.add_chart(pass_chart, "K37")

    # ===== DOWNLOAD =====
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=subject_analysis.xlsx'

    wb.save(response)
    return response

def download_cumulative_backlog(request):
    from django.http import HttpResponse
    from openpyxl import Workbook

    qs = Result.objects.select_related('student','subject')

    wb = Workbook()
    ws = wb.active

    ws.append(["Name","Subject","Marks"])

    for r in qs.filter(marks__lt=35):
        ws.append([r.student.name, r.subject.code, r.marks])

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition']='attachment; filename=backlog.xlsx'
    wb.save(response)
    return response

# ================= CATEGORY =================
def download_category_report(request):
    from django.http import HttpResponse
    from openpyxl import Workbook
    from openpyxl.chart import BarChart, PieChart, Reference
    from django.db.models import Count, Avg, Q

    PASS = 35

    qs = Result.objects.select_related('student')

    year = request.GET.get('year')
    sem = request.GET.get('sem')

    if year:
        qs = qs.filter(year=year)
    if sem:
        qs = qs.filter(semester=sem)

    # ===== AGGREGATE =====
    data = qs.values('student__actual_category').annotate(
        students=Count('student', distinct=True),
        avg=Avg('marks'),
        total=Count('id'),
        passed=Count('id', filter=Q(marks__gte=PASS)),
        failed=Count('id', filter=Q(marks__lt=PASS)),
        distinction=Count('id', filter=Q(marks__gte=75)),
        first=Count('id', filter=Q(marks__gte=60, marks__lt=75)),
        pass_cls=Count('id', filter=Q(marks__gte=35, marks__lt=60)),
        below50=Count('id', filter=Q(marks__lt=50)),
        between50_75=Count('id', filter=Q(marks__gte=50, marks__lt=75)),
        above75=Count('id', filter=Q(marks__gte=75)),
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Category Analysis"

    # ===== TABLE =====
    ws.append([
        "Category","Students","Avg",
        "Pass","Fail",
        "Distinction","First Class","Pass Class",
        "<50","50-75",">75"
    ])

    for d in data:
        ws.append([
            d['student__actual_category'],
            d['students'],
            round(d['avg'] or 0,2),
            d['passed'],
            d['failed'],
            d['distinction'],
            d['first'],
            d['pass_cls'],
            d['below50'],
            d['between50_75'],
            d['above75']
        ])

    last = ws.max_row
    cats = Reference(ws, min_col=1, min_row=2, max_row=last)

    # =========================
    # 1️⃣ PIE (DISTRIBUTION)
    # =========================
    pie = PieChart()
    pie.title = "Category Distribution"

    data_ref = Reference(ws, min_col=2, min_row=1, max_row=last)
    pie.add_data(data_ref, titles_from_data=True)
    pie.set_categories(cats)

    ws.add_chart(pie, "L2")

    # =========================
    # 2️⃣ AVG MARKS
    # =========================
    avg_chart = BarChart()
    avg_chart.title = "Average Marks"

    data_ref = Reference(ws, min_col=3, min_row=1, max_row=last)
    avg_chart.add_data(data_ref, titles_from_data=True)
    avg_chart.set_categories(cats)

    for s in avg_chart.series:
        s.data_labels = True

    ws.add_chart(avg_chart, "L18")

    # =========================
    # 3️⃣ PASS vs FAIL
    # =========================
    pf_chart = BarChart()
    pf_chart.title = "Pass vs Fail"

    data_ref = Reference(ws, min_col=4, max_col=5, min_row=1, max_row=last)
    pf_chart.add_data(data_ref, titles_from_data=True)
    pf_chart.set_categories(cats)

    for s in pf_chart.series:
        s.data_labels = True

    ws.add_chart(pf_chart, "L34")

    # =========================
    # 4️⃣ MULTI BAR (CLASSIFICATION)
    # =========================
    multi = BarChart()
    multi.type = "col"
    multi.grouping = "clustered"
    multi.title = "Result Classification"

    data_ref = Reference(ws, min_col=6, max_col=8, min_row=1, max_row=last)
    multi.add_data(data_ref, titles_from_data=True)
    multi.set_categories(cats)

    multi.gapWidth = 150

    for s in multi.series:
        s.data_labels = True

    ws.add_chart(multi, "L50")

    # =========================
    # 5️⃣ SCORE DISTRIBUTION
    # =========================
    score = BarChart()
    score.type = "col"
    score.grouping = "clustered"
    score.title = "Score Distribution"

    data_ref = Reference(ws, min_col=9, max_col=11, min_row=1, max_row=last)
    score.add_data(data_ref, titles_from_data=True)
    score.set_categories(cats)

    for s in score.series:
        s.data_labels = True

    ws.add_chart(score, "L66")

    # =========================
    # DOWNLOAD
    # =========================
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=category_full.xlsx'

    wb.save(response)
    return response

# ================= BRANCH =================
def download_branch_report(request):
    from django.http import HttpResponse
    from openpyxl import Workbook
    from openpyxl.chart import BarChart, Reference
    from django.db.models import Count, Avg, Q

    PASS = 35

    qs = Result.objects.select_related('student')

    year = request.GET.get('year')
    sem = request.GET.get('sem')

    if year:
        qs = qs.filter(year=year)
    if sem:
        qs = qs.filter(semester=sem)

    branches = qs.values('student__branch').distinct()

    data = []

    for b in branches:
        name = b['student__branch']
        bqs = qs.filter(student__branch=name)

        students = bqs.values('student').distinct().count()
        avg = round(bqs.aggregate(avg=Avg('marks'))['avg'] or 0, 2)

        total = bqs.count()
        passed = bqs.filter(marks__gte=PASS).count()
        failed = total - passed

        pass_pct = round((passed/total)*100,2) if total else 0

        distinction = bqs.filter(marks__gte=75).count()

        data.append({
            'branch': name,
            'students': students,
            'avg': avg,
            'pass_pct': pass_pct,
            'backlogs': failed,
            'distinction': distinction
        })

    # ===== EXCEL =====
    wb = Workbook()
    ws = wb.active
    ws.title = "Branch Analysis"

    ws.append(["Branch","Students","Avg","Pass %","Backlogs","Distinction"])

    for d in data:
        ws.append([
            d['branch'],
            d['students'],
            d['avg'],
            d['pass_pct'],
            d['backlogs'],
            d['distinction']
        ])

    # ===== GRAPH 1: PASS % =====
    chart1 = BarChart()
    chart1.title = "Pass % by Branch"

    data_ref = Reference(ws, min_col=4, min_row=1, max_row=len(data)+1)
    cats_ref = Reference(ws, min_col=1, min_row=2, max_row=len(data)+1)

    chart1.add_data(data_ref, titles_from_data=True)
    chart1.set_categories(cats_ref)

    ws.add_chart(chart1, "H2")

    # ===== GRAPH 2: BACKLOGS =====
    chart2 = BarChart()
    chart2.title = "Backlogs by Branch"

    data_ref2 = Reference(ws, min_col=5, min_row=1, max_row=len(data)+1)

    chart2.add_data(data_ref2, titles_from_data=True)
    chart2.set_categories(cats_ref)

    ws.add_chart(chart2, "H20")

    # ===== DOWNLOAD =====
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=branch_analysis.xlsx'

    wb.save(response)
    return response


# ================= BACKLOG =================
def download_backlog_report(request):
    from django.http import HttpResponse
    from openpyxl import Workbook
    from openpyxl.chart import LineChart, Reference
    from django.db.models import Count, Q
    from collections import defaultdict

    PASS = 35

    qs = Result.objects.select_related('student','subject')

    year = request.GET.get('year')
    sem = request.GET.get('sem')

    if year:
        qs = qs.filter(year=year)
    if sem:
        qs = qs.filter(semester=sem)

    current_backlogs = qs.filter(marks__lt=PASS)

    all_results = Result.objects.select_related('student','subject')

    # ===== TABLE =====
    wb = Workbook()
    ws = wb.active
    ws.title = "Backlog Data"

    ws.append(["USN","Name","Branch","Current","Total","Subjects"])

    student_data = defaultdict(lambda: {
        "name": "",
        "branch": "",
        "current": [],
        "cumulative": []
    })

    for r in current_backlogs:
        usn = r.student.usn
        student_data[usn]["name"] = r.student.name
        student_data[usn]["branch"] = r.student.branch
        student_data[usn]["current"].append(r.subject.code)

    for r in all_results.filter(marks__lt=PASS):
        usn = r.student.usn
        student_data[usn]["cumulative"].append(r.subject.code)

    for usn, d in student_data.items():
        if not d["current"]:
            continue

        unique = list(set(d["cumulative"]))

        ws.append([
            usn,
            d["name"],
            d["branch"],
            ", ".join(d["current"]),
            len(unique),
            ", ".join(unique)
        ])

    # =========================================
    # 🔥 TREND SHEET (IMPORTANT)
    # =========================================
    ws2 = wb.create_sheet("Trend Analysis")

    ws2.append(["Year-Sem","Backlogs"])

    trend = all_results.values('year','semester').annotate(
        backlog_count=Count('id', filter=Q(marks__lt=PASS))
    ).order_by('year','semester')

    # if NO DATA → still show something
    if not trend:
        ws2.append(["No Data", 0])
    else:
        for t in trend:
            label = f"{t['year']}-S{t['semester']}"
            ws2.append([label, t['backlog_count']])

    # =========================================
    # 🔥 LINE CHART
    # =========================================
    if ws2.max_row > 1:
        chart = LineChart()
        chart.title = "Backlog Trend (Year-wise)"

        data = Reference(ws2, min_col=2, min_row=1, max_row=ws2.max_row)
        cats = Reference(ws2, min_col=1, min_row=2, max_row=ws2.max_row)

        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)

        chart.y_axis.title = "Backlogs"
        chart.x_axis.title = "Year-Sem"

        ws2.add_chart(chart, "B5")

    # ===== DOWNLOAD =====
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=backlog_report.xlsx'

    wb.save(response)
    return response
# ================= QUOTA =================
def download_quota_report(request):
    from django.http import HttpResponse
    from openpyxl import Workbook
    from openpyxl.chart import BarChart, PieChart, Reference
    from django.db.models import Count, Avg, Q

    PASS = 35

    qs = Result.objects.select_related('student')

    year = request.GET.get('year')
    sem = request.GET.get('sem')

    if year:
        qs = qs.filter(year=year)
    if sem:
        qs = qs.filter(semester=sem)

    data = qs.values('student__admission_quota').annotate(
        students=Count('student', distinct=True),
        total=Count('id'),
        passed=Count('id', filter=Q(marks__gte=PASS)),
        failed=Count('id', filter=Q(marks__lt=PASS)),
        avg=Avg('marks'),
        below50=Count('id', filter=Q(marks__lt=50)),
        between50_75=Count('id', filter=Q(marks__gte=50, marks__lt=75)),
        above75=Count('id', filter=Q(marks__gte=75)),
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Quota Analysis"

    # ===== TABLE =====
    headers = [
        "Quota","Students","Avg Marks","Pass %","Fail %",
        "Pass Count","Backlog Count","<50","50-75",">=75"
    ]
    ws.append(headers)

    rows = []
    for d in data:
        total = d['total']
        pass_pct = round((d['passed']/total)*100, 2) if total else 0
        fail_pct = round((d['failed']/total)*100, 2) if total else 0

        row = [
            d['student__admission_quota'],
            d['students'],
            round(d['avg'] or 0, 2),
            pass_pct,
            fail_pct,
            d['passed'],
            d['failed'],
            d['below50'],
            d['between50_75'],
            d['above75']
        ]
        rows.append(row)
        ws.append(row)

    last_row = len(rows) + 1

    # ===== CATEGORY LABELS =====
    cats = Reference(ws, min_col=1, min_row=2, max_row=last_row)

    # ===== PIE CHART =====
    pie = PieChart()
    pie.title = "Student Distribution"
    data_ref = Reference(ws, min_col=2, min_row=1, max_row=last_row)
    pie.add_data(data_ref, titles_from_data=True)
    pie.set_categories(cats)
    ws.add_chart(pie, "L2")

    # ===== PASS % vs AVG (CLEAR CLUSTERED) =====
    bar = BarChart()
    bar.type = "col"
    bar.grouping = "clustered"
    bar.title = "Pass % vs Avg"

    data_ref = Reference(ws, min_col=3, max_col=4, min_row=1, max_row=last_row)
    bar.add_data(data_ref, titles_from_data=True)
    bar.set_categories(cats)

    bar.x_axis.tickLblPos = "low"
    bar.gapWidth = 150

    for s in bar.series:
        s.data_labels = True

    ws.add_chart(bar, "L20")

    # ===== PASS vs BACKLOG (CLEAR) =====
    pb = BarChart()
    pb.type = "col"
    pb.grouping = "clustered"
    pb.title = "Pass vs Backlog"

    data_ref = Reference(ws, min_col=6, max_col=7, min_row=1, max_row=last_row)
    pb.add_data(data_ref, titles_from_data=True)
    pb.set_categories(cats)

    pb.x_axis.tickLblPos = "low"
    pb.gapWidth = 150

    for s in pb.series:
        s.data_labels = True

    ws.add_chart(pb, "L38")

    # ===== SCORE DISTRIBUTION =====
    score = BarChart()
    score.type = "col"
    score.grouping = "clustered"
    score.title = "Score Distribution"

    data_ref = Reference(ws, min_col=8, max_col=10, min_row=1, max_row=last_row)
    score.add_data(data_ref, titles_from_data=True)
    score.set_categories(cats)

    score.x_axis.tickLblPos = "low"
    score.gapWidth = 150

    for s in score.series:
        s.data_labels = True

    ws.add_chart(score, "L56")
    

    # ===== DOWNLOAD =====
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=quota_analysis.xlsx'

    wb.save(response)
    return response

def download_report(request):

    report_type = request.GET.get('type')

    if report_type == "subject":
        return download_subject_report(request)

    elif report_type == "branch":
        return download_branch_report(request)

    elif report_type == "category":
        return download_category_report(request)

    elif report_type == "quota":
        return download_quota_report(request)

    elif report_type == "backlog":
        return download_backlog_report(request)

    elif report_type == "dashboard":
        return download_dashboard_report(request)   # 🔥 ADD THIS

    else:
        return download_subject_report(request)  # fallback
    
import json, base64
from io import BytesIO
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from django.db.models import Avg, Q

def download_report_excel(request):
    from .models import Result   # adjust if needed

    PASS = 35

    qs = Result.objects.select_related('student')

    year = request.GET.get('year')
    sem = request.GET.get('sem')

    if year:
        qs = qs.filter(year=year)
    if sem:
        qs = qs.filter(semester=sem)

    # ===== BUILD SAME TABLE DATA =====
    branches = qs.values('student__branch').distinct()

    data = []

    for b in branches:
        name = b['student__branch']
        bqs = qs.filter(student__branch=name)

        students = bqs.values('student').distinct().count()
        avg = round(bqs.aggregate(avg=Avg('marks'))['avg'] or 0, 2)

        total = bqs.count()
        passed = bqs.filter(marks__gte=PASS).count()
        failed = total - passed

        pass_pct = round((passed/total)*100,2) if total else 0
        distinction = bqs.filter(marks__gte=75).count()

        data.append([name, students, avg, pass_pct, failed, distinction])

    # ===== GET CHART IMAGES =====
    req_data = json.loads(request.body)
    images = req_data.get('images', [])

    wb = Workbook()
    ws = wb.active
    ws.title = "Branch Analysis"

    # ===== ADD TABLE =====
    ws.append(["Branch","Students","Avg","Pass %","Backlogs","Distinction"])

    for row in data:
        ws.append(row)

    # ===== ADD CHARTS BELOW TABLE =====
    start_row = len(data) + 4

    for img_data in images:
        img_bytes = base64.b64decode(img_data.split(',')[1])
        img = Image(BytesIO(img_bytes))

        ws.add_image(img, f"A{start_row}")
        start_row += 22

    # ===== RESPONSE =====
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=branch_analysis.xlsx'

    wb.save(response)
    return response

def single_student_cumulative_backlog(request):
    from django.http import HttpResponse
    from openpyxl import Workbook
    from openpyxl.chart import BarChart, LineChart, Reference
    from django.db.models import Count, Q
    from collections import defaultdict

    PASS = 35

    qs = Result.objects.select_related('student','subject')

    year = request.GET.get('year')
    sem = request.GET.get('sem')

    if year:
        qs = qs.filter(year=year)
    if sem:
        qs = qs.filter(semester=sem)

    current_backlogs = qs.filter(marks__lt=PASS)
    all_results = Result.objects.select_related('student','subject')

    # =====================
    # DATA COLLECTION
    # =====================
    student_data = defaultdict(int)
    branch_data = defaultdict(int)
    severity_data = defaultdict(int)

    for r in current_backlogs:
        student_data[r.student.usn] += 1
        branch_data[r.student.branch] += 1
        severity_data[r.student.usn] += 1

    # =====================
    # EXCEL
    # =====================
    wb = Workbook()

    # ===== SHEET 1: TABLE =====
    ws = wb.active
    ws.title = "Backlog Data"

    ws.append(["USN","Backlogs"])

    for usn, count in student_data.items():
        ws.append([usn, count])

    ws.append([])
    ws.append(["Branch","Backlogs"])

    for b, count in branch_data.items():
        ws.append([b, count])

    # =====================
    # SHEET 2: CHARTS
    # =====================
    ws2 = wb.create_sheet("Charts")

    # =====================
    # CHART 1: TOP STUDENTS
    # =====================
    ws2.append(["USN","Backlogs"])
    for usn, count in student_data.items():
        ws2.append([usn, count])

    chart1 = BarChart()
    chart1.title = "Top Backlog Students"

    data = Reference(ws2, min_col=2, min_row=1, max_row=len(student_data)+1)
    cats = Reference(ws2, min_col=1, min_row=2, max_row=len(student_data)+1)

    chart1.add_data(data, titles_from_data=True)
    chart1.set_categories(cats)

    ws2.add_chart(chart1, "A10")

    # =====================
    # CHART 2: BRANCH
    # =====================
    start_row = ws2.max_row + 3

    ws2.cell(row=start_row, column=1, value="Branch")
    ws2.cell(row=start_row, column=2, value="Backlogs")

    row = start_row + 1
    for b, count in branch_data.items():
        ws2.cell(row=row, column=1, value=b)
        ws2.cell(row=row, column=2, value=count)
        row += 1

    chart2 = BarChart()
    chart2.title = "Branch-wise Backlogs"

    data = Reference(ws2, min_col=2, min_row=start_row, max_row=row-1)
    cats = Reference(ws2, min_col=1, min_row=start_row+1, max_row=row-1)

    chart2.add_data(data, titles_from_data=True)
    chart2.set_categories(cats)

    ws2.add_chart(chart2, "A30")

    # =====================
    # CHART 3: SEVERITY
    # =====================
    start_row = ws2.max_row + 3

    ws2.cell(row=start_row, column=1, value="Severity")
    ws2.cell(row=start_row, column=2, value="Count")

    ws2.cell(row=start_row+1, column=1, value="Total")
    ws2.cell(row=start_row+1, column=2, value=sum(severity_data.values()))

    chart3 = BarChart()
    chart3.title = "Backlog Severity"

    data = Reference(ws2, min_col=2, min_row=start_row, max_row=start_row+1)
    cats = Reference(ws2, min_col=1, min_row=start_row+1, max_row=start_row+1)

    chart3.add_data(data, titles_from_data=True)
    chart3.set_categories(cats)

    ws2.add_chart(chart3, "A50")

    # =====================
    # 🔥 CHART 4: TREND (NEW)
    # =====================
    start_row = ws2.max_row + 5

    ws2.cell(row=start_row, column=1, value="Year-Sem")
    ws2.cell(row=start_row, column=2, value="Backlogs")

    trend = Result.objects.values('year','semester').annotate(
        backlog_count=Count('id', filter=Q(marks__lt=PASS))
    ).order_by('year','semester')

    row = start_row + 1
    for t in trend:
        label = f"{t['year']}-S{t['semester']}"
        ws2.cell(row=row, column=1, value=label)
        ws2.cell(row=row, column=2, value=t['backlog_count'])
        row += 1

    if row > start_row + 1:
        chart4 = LineChart()
        chart4.title = "Backlog Trend (Year-Sem)"

        data = Reference(ws2, min_col=2, min_row=start_row, max_row=row-1)
        cats = Reference(ws2, min_col=1, min_row=start_row+1, max_row=row-1)

        chart4.add_data(data, titles_from_data=True)
        chart4.set_categories(cats)

        chart4.y_axis.title = "Backlogs"
        chart4.x_axis.title = "Year-Sem"

        ws2.add_chart(chart4, "A70")

    # =====================
    # DOWNLOAD
    # =====================
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=backlog_analysis.xlsx'

    wb.save(response)
    return response